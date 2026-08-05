import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def generer_reporting(df_detail, df_site, df_marque):
    """Génère un fichier Excel complet avec plusieurs onglets"""

    chemin = "C:/Users/florian/pandas_project/data/output/reporting_carbase.xlsx"
    mois   = datetime.now().strftime("%B %Y")

    # Écrire les 3 onglets
    with pd.ExcelWriter(chemin, engine="openpyxl") as writer:
        df_detail.to_excel(writer, sheet_name="Détail",      index=False)
        df_site.to_excel(writer,   sheet_name="Par Site",    index=False)
        df_marque.to_excel(writer, sheet_name="Par Marque",  index=False)

    # Mise en forme
    wb = load_workbook(chemin)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    green_fill  = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red_fill    = PatternFill(fill_type="solid", fgColor="FFC7CE")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # En-tête
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")

        # Largeur des colonnes
        for col in ws.columns:
            max_width = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_width + 4

        # Mise en forme conditionnelle sur ecart_pct
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column_letter in ["F", "G"] and isinstance(cell.value, float):
                    if "ecart_pct" in str(ws.cell(1, cell.column).value or "").lower():
                        cell.fill = green_fill if cell.value > 0 else red_fill

    wb.save(chemin)
    print(f"✅ Reporting Excel généré — {chemin}")
    return chemin

if __name__ == "__main__":
    import sys
    import os
    sys.path.append(os.path.dirname(__file__))  # ✅ ajoute le dossier pipeline au chemin
    
    from transformation import transformer_donnees, agreger_par_site, agreger_par_marque
    
    df = pd.read_excel("C:/Users/florian/pandas_project/data/input/export_carbase.xlsx")
    df_transforme = transformer_donnees(df)
    df_site       = agreger_par_site(df_transforme)
    df_marque     = agreger_par_marque(df_transforme)
    generer_reporting(df_transforme, df_site, df_marque)